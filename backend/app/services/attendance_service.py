import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os
from app.core.config import settings
from app.models.student import Student

class AttendanceService:
    def __init__(self):
        self.file_path = settings.ATTENDANCE_FILE
        self._ensure_file()

    def _ensure_file(self):
        if not os.path.exists(self.file_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Student ID", "Program", "Major", "Date", "Time", "Status"])
            wb.save(self.file_path)

    def log_attendance(self, student: Student):
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
        except FileNotFoundError:
            self._ensure_file()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
        
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        
        ws.append([student.name, student.student_id, student.program, student.major, date_str, time_str, "Present"])
        wb.save(self.file_path)

    def get_attendance_records(self):
        if not os.path.exists(self.file_path):
            return []
        
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        records = []
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Ensure row has data
                records.append({
                    "name": row[0],
                    "student_id": row[1],
                    "program": row[2],
                    "major": row[3],
                    "date": row[4],
                    "time": row[5],
                    "status": row[6]
                })
        # Return reversed list to show newest first
        return records[::-1]
